
import openpyxl
from openpyxl import load_workbook
import os

class Timing:
    """
    Class for representing timing data.
    Will accept inputs from the following sources:
        - Excel workbook
    """

    def __init__(self, input=None):
        """

        """

        self.signals = dict()           # capture all signal waveform information (raw)
        self.configs = dict()           # capture information that may influence how data is displayed/interpreted
        self.comments = dict()          # capture misc comments
        self.colors = dict()            # colormap of different waveforms
        self.colors['blank'] = '00000000'

        self.config_keywords = [
            'pw_break',                    # assign the color in the cell following this keyword to be identified as a break
            'pw_x',                        # assign the color in the cell following this keyword to be an 'x'
            'pw_pclk',                     # assign the color in the cell following this keyword to be a posedge clock
            'pw_nclk',                     # assign the color in the cell following this keyword to be a posedge clock
            'pw_ignore'                    # assign the color in the cell following this keyword to be ignored (do not process)
        ]

        if input is not None:
            file_path, file_extension = os.path.splittext(input)
            if file_extension == ".xls" or file_extension == ".xlsx":
                self.read_xls(input)

    def read_xls_header(self, worksheet):
        """
        Read the header line of an excel file to get column configurations
        - name: signal name
        - group: signal grouping (if any)
        - type:
        """

        header_keywords = ["name", "group"]
        self.xls_header = {
            'name': 0,
            'edge': None,
            'group': None
        }

        start_col = 0
        for col in range(0, worksheet.max_column):
            cell = worksheet[1][col]
            if str(cell.value).lower() in header_keywords:
                self.xls_header[str(cell.value).lower()] = col
                start_col = col
        self.xls_header['start_col'] = start_col + 1

    def read_xls(self, xlfile, sheet=None, header=True):
        """
        Read an Excel workbook
        """

        wb = load_workbook(xlfile)
        if sheet is None:               # Grab the first sheet if none is specified
            sheet = wb.sheetnames[0]
        sh = wb[sheet]
        self.sh = sh
        self.wb = wb

        if header == True:
            start_row = 2
            self.read_xls_header(worksheet=sh)
            name_idx = self.xls_header['name']
            edge_idx = self.xls_header['edge']
            group_idx = self.xls_header['group']
            start_col = self.xls_header['start_col']
        else:
            start_row = 1
            name_idx = 0
            edge_idx = None
            group_idx = None
            start_col = 1

        for row in range(start_row, sh.max_row+1):
            sig_name = sh[row][name_idx].value
            if group_idx is not None:
                group_name = sh[row][group_idx].value
            if sig_name is not None:
                # Check if the line is a configuration keyword
                if sig_name in self.config_keywords:
                    color_in_hex = sh[row][name_idx].fill.start_color.index
                    if color_in_hex != self.colors['blank']:  # If it's not just a blank cell (whatever that's defined as) assign the color to be a break, ignore, etc...
                        self.configs[color_in_hex] = sig_name
                # Check if this is a new signal name (avoid duplicates -- first will be used)
                elif sig_name not in self.signals.keys():
                    self.signals[sig_name] = dict()
                    self.signals[sig_name]['name'] = sig_name
                    raw_data = []
                    wave = ""
                    color_in_hex = self.colors['blank']
                    for col in range(start_col, sh.max_column):
                        cell = sh[row][col]
                        last_color = color_in_hex
                        color_in_hex = cell.fill.start_color.index
                        # Check for keyword colors
                        if color_in_hex in self.configs.keys():
                            if self.configs[color_in_hex] == 'pw_break':
                                wave += "|"
                            if self.configs[color_in_hex] == 'pw_pclk':
                                wave += "p"
                            color_in_hex = last_color                   # When we have config patterns, set the data back to the previous value
                        # The cell is filled with some color that is not a keyword
                        elif color_in_hex != self.colors['blank']:      
                            value = cell.value
                            if cell.value is None:
                                if (len(wave) > 0) and (color_in_hex == last_color):
                                    wave += "."
                                else:
                                    wave += "1"
                            else:
                                color = self.map_color(color_in_hex)
                                raw_data.append(value)
                                wave += color
                        else:                               # The cell is blank (should be 0)
                            if (len(wave) > 0) and (color_in_hex == last_color):
                                wave += "."
                            else:
                                wave += "0"

                    self.signals[sig_name]["data"] = raw_data
                    self.signals[sig_name]["wave"] = wave

    def map_color(self, color_in_hex):
        """
        Maps each unique color found in the xls spreadsheet to a wavedrom color index item
        """
        if color_in_hex not in self.colors.keys():
            self.colors[color_in_hex] = str(2+len(self.colors)%8)   # Assign colors in order of appearance
        return self.colors[color_in_hex]

    def get_diagram(self, signal_list = None):
        """
        signal_list: list of signal names, in order, that we wish to draw a diagram for.
        """
        diagram = dict()
        diagram['signal'] = []
        if signal_list is None:
            for sig in self.signals.keys():
                try:
                    diagram['signal'].append(self.signals[sig])
                except:
                    pass
        else:
            for sig in signal_list:
                try:
                    diagram['signal'].append(self.signals[sig])
                except:
                    pass
        return diagram